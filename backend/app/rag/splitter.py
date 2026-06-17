"""Document splitters for multiple formats: Markdown, FAQ, PDF, Word, Excel, plain text."""

from __future__ import annotations

import re
from dataclasses import dataclass, field


@dataclass
class Chunk:
    """A chunk unit returned by any splitter pipeline."""
    content: str
    chunk_index: int
    metadata: dict[str, str] = field(default_factory=dict)


# ── Text splitters ────────────────────────────────────────────────────────────

class MarkdownSplitter:
    """Split Markdown by H1/H2/H3 structure."""

    def split(self, text: str) -> list[Chunk]:
        chunks: list[Chunk] = []
        sections = re.split(r"(?=^#{1,3}\s)", text, flags=re.MULTILINE)
        idx = 0
        for section in sections:
            section = section.strip()
            if not section:
                continue
            heading_match = re.match(r"^(#{1,3})\s+(.+)", section)
            heading = heading_match.group(2) if heading_match else ""
            chunks.append(Chunk(content=section, chunk_index=idx, metadata={"heading": heading}))
            idx += 1
        return chunks


class FAQSplitter:
    """Split by Q&A natural boundaries."""

    def split(self, text: str) -> list[Chunk]:
        pairs = re.split(r"(?=^(?:Q|问)[：:])", text, flags=re.MULTILINE)
        chunks: list[Chunk] = []
        idx = 0
        for pair in pairs:
            pair = pair.strip()
            if not pair:
                continue
            chunks.append(Chunk(content=pair, chunk_index=idx))
            idx += 1
        return chunks


class FixedSizeSplitter:
    """Fixed-size sliding window splitter for plain text."""

    def __init__(self, chunk_size: int = 512, overlap: int = 64) -> None:
        self.chunk_size = chunk_size
        self.overlap = overlap

    def split(self, text: str) -> list[Chunk]:
        chunks: list[Chunk] = []
        idx, start = 0, 0
        while start < len(text):
            end = min(start + self.chunk_size, len(text))
            chunks.append(Chunk(content=text[start:end], chunk_index=idx))
            idx += 1
            if end >= len(text):
                break
            start = end - self.overlap
        return chunks


class SemanticSplitter:
    """Split by paragraph/section boundaries for documents with structure."""

    def __init__(self, max_chunk_size: int = 800, min_chunk_size: int = 200) -> None:
        self.max_chunk_size = max_chunk_size
        self.min_chunk_size = min_chunk_size

    def split(self, text: str) -> list[Chunk]:
        # First, split by double newlines (paragraphs)
        paras = [p.strip() for p in re.split(r'\n\s*\n', text) if p.strip()]

        chunks: list[Chunk] = []
        idx = 0
        current = ""
        for para in paras:
            if len(current) + len(para) < self.max_chunk_size:
                current = (current + "\n\n" + para).strip()
            else:
                if len(current) >= self.min_chunk_size:
                    chunks.append(Chunk(content=current, chunk_index=idx))
                    idx += 1
                    current = para
                else:
                    current = (current + "\n\n" + para).strip()

        if current.strip():
            chunks.append(Chunk(content=current, chunk_index=idx))

        return chunks


class ExcelSplitter:
    """Split Excel by rows, preserving header row in each chunk."""

    def __init__(self, rows_per_chunk: int = 10) -> None:
        self.rows_per_chunk = rows_per_chunk

    def split(self, rows: list[list[str]], headers: list[str] | None = None) -> list[Chunk]:
        if not rows:
            return []

        header_line = "\t".join(headers) if headers else ""
        chunks: list[Chunk] = []
        idx = 0
        for i in range(0, len(rows), self.rows_per_chunk):
            batch = rows[i:i + self.rows_per_chunk]
            lines = [header_line] if header_line else []
            for row in batch:
                lines.append("\t".join(str(c) for c in row))
            chunks.append(Chunk(
                content="\n".join(lines),
                chunk_index=idx,
                metadata={"rows": f"{i}-{i + len(batch)}", "format": "table"},
            ))
            idx += 1
        return chunks


# ── Text extraction ───────────────────────────────────────────────────────────

def extract_text(file_path: str, file_type: str) -> str:
    """Extract plain text from PDF, Word, or Excel files."""
    if file_type == "pdf":
        return _extract_pdf(file_path)
    elif file_type == "docx":
        return _extract_docx(file_path)
    elif file_type == "xlsx":
        return _extract_xlsx(file_path)
    return ""


def _extract_pdf(path: str) -> str:
    """Extract text from PDF using PyPDF2 or pdfplumber."""
    try:
        import pdfplumber
        texts: list[str] = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texts.append(t)
        return "\n\n".join(texts)
    except ImportError:
        pass

    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(path)
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)
    except ImportError:
        raise ImportError("需要安装 pdfplumber 或 PyPDF2: pip install pdfplumber")


def _extract_docx(path: str) -> str:
    """Extract text from Word .docx using python-docx."""
    try:
        from docx import Document
        doc = Document(path)
        paras: list[str] = []
        for para in doc.paragraphs:
            if para.style.name.startswith("Heading"):
                level = para.style.name.split()[-1]
                try:
                    lv = int(level)
                except ValueError:
                    lv = 1
                paras.append(f"{'#' * min(lv, 3)} {para.text}")
            elif para.text.strip():
                paras.append(para.text)
        return "\n\n".join(paras)
    except ImportError:
        raise ImportError("需要安装 python-docx: pip install python-docx")


def _extract_xlsx(path: str) -> str:
    """Extract text from Excel, returning tab-separated rows."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        all_lines: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_lines.append(f"## {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                values = [str(c) if c is not None else "" for c in row]
                all_lines.append("\t".join(values))
        wb.close()
        return "\n".join(all_lines)
    except ImportError:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")


# ── Routing ───────────────────────────────────────────────────────────────────

def auto_split(text: str, file_type: str, file_path: str = "") -> list[Chunk]:
    """Route to the correct splitter based on file type."""
    if file_type == "md":
        return MarkdownSplitter().split(text)
    if file_type == "faq":
        return FAQSplitter().split(text)
    if file_type in ("pdf", "docx"):
        return SemanticSplitter().split(text)
    if file_type == "xlsx":
        # Re-parse the tab-separated text back into rows for ExcelSplitter
        lines = text.strip().split("\n")
        rows: list[list[str]] = []
        headers: list[str] = []
        for i, line in enumerate(lines):
            if line.startswith("## "):
                continue  # skip sheet name markers
            cells = line.split("\t")
            if i == 0:
                headers = cells
            else:
                rows.append(cells)
        return ExcelSplitter(rows_per_chunk=10).split(rows, headers)
    return FixedSizeSplitter().split(text)
